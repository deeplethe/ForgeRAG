"""
Spreadsheet-as-document backend.

Native parser for ``.xlsx`` / ``.csv`` / ``.tsv`` uploads. Produces a
ParsedDocument where each sheet becomes one ``BlockType.TABLE`` block
(CSV / TSV always have one sheet; xlsx may have many).

Design (see ``docs/roadmaps/spreadsheet-as-document.md`` for the
full rationale):

  * Each sheet → 1 TABLE block.
  * ``block.text`` is the **description** (filled by the
    ``table_enrichment`` phase later in the pipeline). For now this
    backend leaves it empty — same pattern as image-as-document
    where the VLM enrichment phase fills the text.
  * ``block.table_markdown`` carries the full sheet rendered as a
    GFM markdown table. This is **not** part of the retrieval
    index; it lives there for future consumers (the SpreadsheetViewer
    component, a future ``query_spreadsheet`` agent tool).
  * ``block.bbox = (0,0,0,0)`` sentinel — there is no spatial
    layout to highlight on a sheet.
  * Sheet name lives in two places:
      - ``Page.name`` — structured access for the frontend tab
        strip.
      - injected as the heading of ``block.text`` (after the
        enrichment phase fills it) so retrieval naturally sees the
        sheet name without parsing the structured field.

Edge cases handled at parse time:

  * ``.xls`` (legacy binary) is rejected before reaching here by
    ``LEGACY_OFFICE_EXTENSIONS`` at the upload route. We don't
    handle it.
  * Multi-sheet xlsx → each sheet a Page + a TABLE block; ``page_no``
    is ``sheet_index + 1``.
  * Empty sheet → block has no rows; markdown is just the header
    line. Description phase still describes it ("0 data rows").
  * Cell-count > ``SPREADSHEET_MAX_CELLS`` is rejected at the
    upload route, not here. This backend trusts the gate has run.
  * CSV encoding is auto-detected via ``charset_normalizer``.
  * Headers detected heuristically — if the first row looks like
    data (numeric / boolean), we synthesise ``col_1, col_2, ...``
    column names rather than promote the data row.
"""

from __future__ import annotations

import csv
import io
import logging
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from ..blob_store import BlobStore
from ..schema import (
    Block,
    BlockType,
    DocFormat,
    DocProfile,
    Page,
    ParsedDocument,
    ParseTrace,
)
from .base import BackendUnavailable, ParserBackend

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public types
# ---------------------------------------------------------------------------


class _SheetData:
    """Lightweight container for one sheet's rows + metadata."""

    __slots__ = ("col_count", "data_rows", "headers", "name", "row_count")

    def __init__(
        self,
        name: str,
        headers: list[str],
        data_rows: list[list[Any]],
    ):
        self.name = name
        self.headers = headers
        self.data_rows = data_rows
        self.row_count = len(data_rows)
        self.col_count = len(headers)


# ---------------------------------------------------------------------------
# CSV / TSV reading
# ---------------------------------------------------------------------------


def _detect_encoding(path: Path) -> str:
    """Best-effort encoding detection for text-format spreadsheets.

    Reads up to 256 KB of the file (enough for confident detection
    on typical CSVs). Falls back to utf-8-sig (handles BOM and plain
    utf-8 transparently) if charset_normalizer can't make a call.
    """
    try:
        from charset_normalizer import from_path

        result = from_path(str(path), steps=5).best()
        if result is not None and result.encoding:
            return result.encoding
    except Exception as e:
        log.debug("charset_normalizer failed for %s: %s", path.name, e)
    return "utf-8-sig"


def _read_csv_rows(path: Path, delimiter: str) -> list[list[str]]:
    """Read a CSV/TSV file → list of row-cell-lists.

    Encoding auto-detected. Streams via ``csv.reader`` so we don't
    materialise the whole file as one Python str — important for
    large CSVs (a 100 MB file could exceed memory if naively read).
    """
    encoding = _detect_encoding(path)
    log.info("CSV %s detected encoding: %s", path.name, encoding)
    rows: list[list[str]] = []
    with path.open("r", encoding=encoding, newline="", errors="replace") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        for row in reader:
            rows.append(row)
    return rows


def _looks_like_data_row(row: list[str]) -> bool:
    """Heuristic: does this row look like data rather than a header?

    Headers are typically short text (column names like "Region" /
    "Q3-Revenue"). Data rows are typically numeric or mixed. We
    flag a row as data if **most** cells parse as numbers.
    """
    if not row:
        return False
    numeric_count = 0
    for cell in row:
        s = (cell or "").strip()
        if not s:
            continue
        try:
            float(s.replace(",", ""))
            numeric_count += 1
        except ValueError:
            pass
    return numeric_count >= max(1, len(row) // 2)


def _split_header(rows: list[list[Any]]) -> tuple[list[str], list[list[Any]]]:
    """Promote first row to headers unless it looks like data.

    Returns ``(headers, data_rows)``. When the first row is
    ambiguous (all-empty, mixed types) we still treat it as a
    header — matches Excel's default behavior on save.
    """
    if not rows:
        return [], []
    first = [str(c) if c is not None else "" for c in rows[0]]
    if _looks_like_data_row(first):
        # Synthesise generic column names so retrieval has something
        # to anchor on (better than blank columns in the markdown).
        headers = [f"col_{i+1}" for i in range(len(first))]
        return headers, [[c for c in r] for r in rows]
    return first, [[c for c in r] for r in rows[1:]]


def _read_csv_or_tsv(path: Path) -> _SheetData:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    raw_rows = _read_csv_rows(path, delimiter)
    headers, data_rows = _split_header(raw_rows)
    return _SheetData(name=path.stem, headers=headers, data_rows=data_rows)


# ---------------------------------------------------------------------------
# XLSX reading
# ---------------------------------------------------------------------------


def _read_xlsx(path: Path) -> list[_SheetData]:
    """Open xlsx in ``read_only=True`` mode (memory-bounded streaming).

    Returns one ``_SheetData`` per sheet, in the order they appear
    in the workbook.

    ``data_only=True`` returns computed values for cells with
    formulas. If the workbook was saved without recalc (rare but
    possible — e.g. saved by a non-Excel tool), the cached value
    may be ``None`` for formula cells.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise BackendUnavailable("openpyxl required for spreadsheet parsing") from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheets: list[_SheetData] = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            headers, data_rows = _split_header(rows)
            sheets.append(
                _SheetData(name=ws.title, headers=headers, data_rows=data_rows)
            )
        return sheets
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Markdown rendering
# ---------------------------------------------------------------------------


def _render_markdown(sheet: _SheetData, max_rows: int | None = None) -> str:
    """Render a sheet to a GFM markdown table.

    ``max_rows=None`` means render everything. Pass a number to cap
    rendering (used by the viewer to avoid massive DOM trees, but
    not by the storage path — we keep the full markdown in
    ``block.table_markdown`` so future agents have access to all
    rows).
    """

    def _escape(cell: Any) -> str:
        s = "" if cell is None else str(cell)
        # Pipe + newline are the GFM table escape pair. Tabs preserved
        # since GFM doesn't care about whitespace inside a cell.
        return s.replace("\\", "\\\\").replace("|", "\\|").replace("\n", " ")

    if not sheet.headers and not sheet.data_rows:
        return ""

    headers = sheet.headers or [f"col_{i+1}" for i in range(sheet.col_count)]
    out = io.StringIO()
    out.write("| " + " | ".join(_escape(h) for h in headers) + " |\n")
    out.write("|" + "|".join(["---"] * len(headers)) + "|\n")
    rows_to_write = (
        sheet.data_rows if max_rows is None else sheet.data_rows[:max_rows]
    )
    for row in rows_to_write:
        # Pad short rows so the markdown stays a valid table even
        # when the source has irregular row widths (common in CSVs
        # exported from older systems).
        cells = list(row) + [""] * (len(headers) - len(row))
        out.write("| " + " | ".join(_escape(c) for c in cells[: len(headers)]) + " |\n")
    return out.getvalue()


def _basic_description(sheet: _SheetData, doc_filename: str, sheet_position: tuple[int, int]) -> str:
    """Deterministic fallback description used when table_enrichment
    is disabled. Pure metadata, no LLM call. Includes enough
    surface area (sheet name, columns, row count) to be retrievable
    via BM25 / vector search even without the LLM-summary layer.
    """
    sheet_idx, sheet_total = sheet_position
    cols = ", ".join(sheet.headers) if sheet.headers else "(no header row)"
    return (
        f"## Sheet: {sheet.name}\n\n"
        f"Columns ({sheet.col_count}): {cols}\n"
        f"Row count: {sheet.row_count}\n"
        f"Source: {doc_filename}, sheet {sheet_idx} of {sheet_total}"
    )


# ---------------------------------------------------------------------------
# Cell counting (used by the upload-route gate)
# ---------------------------------------------------------------------------


def count_cells(path: str | Path) -> int:
    """Cheap pre-flight cell count for the upload-route hard limit.

    For xlsx: opens in ``read_only=True`` and sums
    ``ws.max_row * ws.max_column`` per sheet. ``max_row`` /
    ``max_column`` are O(1) on a read-only workbook (cached on
    sheet metadata) so this doesn't materialise rows.

    For csv/tsv: counts rows by scanning the file and multiplies
    by the column count of the first row. Streams; bounded
    memory.
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(p), read_only=True, data_only=True)
        try:
            total = 0
            for ws in wb.worksheets:
                rows = ws.max_row or 0
                cols = ws.max_column or 0
                total += rows * cols
            return total
        finally:
            wb.close()
    if ext in (".csv", ".tsv"):
        delimiter = "\t" if ext == ".tsv" else ","
        encoding = _detect_encoding(p)
        n_rows = 0
        n_cols = 0
        with p.open("r", encoding=encoding, newline="", errors="replace") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            for i, row in enumerate(reader):
                if i == 0:
                    n_cols = len(row)
                n_rows += 1
        return n_rows * n_cols
    return 0


# ---------------------------------------------------------------------------
# Backend
# ---------------------------------------------------------------------------


class SpreadsheetBackend(ParserBackend):
    """Parses xlsx / csv / tsv into a ParsedDocument with one TABLE
    block per sheet."""

    name = "spreadsheet"

    def __init__(self, blob_store: BlobStore):
        super().__init__(blob_store)

    def parse(
        self,
        path: str,
        doc_id: str,
        parse_version: int,
        profile: DocProfile,
    ) -> ParsedDocument:
        src = Path(path)
        ext = src.suffix.lower()

        if ext == ".xlsx":
            sheets = _read_xlsx(src)
        elif ext in (".csv", ".tsv"):
            sheets = [_read_csv_or_tsv(src)]
        else:
            raise BackendUnavailable(
                f"SpreadsheetBackend can't parse {ext!r}; "
                f"only .xlsx / .csv / .tsv are supported."
            )

        if not sheets:
            raise BackendUnavailable(f"{src.name}: no sheets found")

        blocks: list[Block] = []
        pages: list[Page] = []
        sheet_total = len(sheets)

        for sheet_idx, sheet in enumerate(sheets, start=1):
            page_no = sheet_idx
            block_id = f"{doc_id}:{parse_version}:{page_no}:0"

            # Render the full markdown table once. This is the
            # data-on-the-side that future agents / the viewer
            # consume; not embedded, not searched, just preserved.
            full_md = _render_markdown(sheet, max_rows=None)

            # Initial block.text is the deterministic fallback —
            # the table_enrichment phase will overwrite this with
            # an LLM-generated description if enrichment is
            # configured. If it's not configured, the upload route
            # rejects the upload anyway, so this fallback only ever
            # surfaces in the rare path where someone bypasses the
            # gate (or in tests).
            initial_text = _basic_description(
                sheet,
                doc_filename=src.name,
                sheet_position=(sheet_idx, sheet_total),
            )

            block = Block(
                block_id=block_id,
                doc_id=doc_id,
                parse_version=parse_version,
                page_no=page_no,
                seq=0,
                bbox=(0.0, 0.0, 0.0, 0.0),
                type=BlockType.TABLE,
                text=initial_text,
                table_markdown=full_md,
                table_html=None,
            )
            blocks.append(block)
            pages.append(
                Page(
                    page_no=page_no,
                    width=0.0,
                    height=0.0,
                    name=sheet.name,
                    block_ids=[block_id],
                )
            )

        return ParsedDocument(
            doc_id=doc_id,
            filename=src.name,
            format=DocFormat.SPREADSHEET,
            parse_version=parse_version,
            profile=profile,
            parse_trace=ParseTrace(),  # pipeline fills backend + duration
            pages=pages,
            blocks=blocks,
            toc=None,
        )


# ---------------------------------------------------------------------------
# Helper exposed for tests (and the table_enrichment phase later)
# ---------------------------------------------------------------------------


def split_table_into_row_groups(
    table_markdown: str, rows_per_group: int = 200
) -> list[str]:
    """Split a rendered markdown table into row-group fragments.

    Each fragment is itself a valid markdown table (column header
    repeated at the top of each group). Used by the
    ``table_enrichment`` phase to feed
    ``graph.summarize.summarize_descriptions(kind="table", ...)``
    when the full table doesn't fit one LLM call.

    Fragment shape:
      ```
      | col1 | col2 |
      |------|------|
      | r0   | ...  |
      | r1   | ...  |
      | ... up to rows_per_group rows ...
      ```
    """
    if not table_markdown.strip():
        return []
    lines = table_markdown.splitlines()
    if len(lines) < 3:
        # Only a header (with separator) and no data — return as-is.
        return [table_markdown]
    header, separator = lines[0], lines[1]
    data_lines = lines[2:]
    if len(data_lines) <= rows_per_group:
        return [table_markdown]
    fragments: list[str] = []
    for start in range(0, len(data_lines), rows_per_group):
        chunk = data_lines[start : start + rows_per_group]
        fragments.append("\n".join([header, separator, *chunk]))
    return fragments


def iter_data_rows(sheet: _SheetData) -> Iterable[list[Any]]:
    """Iterate the sheet's data rows. Exposed for tests."""
    yield from sheet.data_rows
